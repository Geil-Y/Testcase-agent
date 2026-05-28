"""CLI for evaluation and legacy optimization helpers.

The old batch generation command has been removed. Use
``python -m testcase_agent.review_pipeline.cli`` for new clarification-first generation runs.
Existing report/evaluation helpers remain for reading completed rounds.
"""

from __future__ import annotations

import argparse
import json
import random
import shutil
import sys
from pathlib import Path

# Add project root to path so we can import testcase_agent
_PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_PROJECT_ROOT))

from openpyxl import load_workbook  # noqa: E402

from testcase_agent.config import get_settings  # noqa: E402
from testcase_agent.parser.html_parser import parse_generated_case  # noqa: E402
from testcase_agent.pipeline.generate import RequirementInput, regenerate_case, run_pipeline  # noqa: E402
from testcase_agent.pipeline.post_process import sanitize_numeric_values  # noqa: E402
from testcase_agent.prompts import render_prompt  # noqa: E402
from testcase_agent.provider.factory import create_provider  # noqa: E402
from testcase_agent.quality.gate import evaluate_cases  # noqa: E402
from optimization.evaluator import CHECKLIST, evaluate_case  # noqa: E402

_VALID_CATEGORIES = {"signal", "threshold", "timing", "state", "observation"}


def _case_to_dict(case) -> dict:
    """Convert a GeneratedCase dataclass to a dict for evaluate_case()."""
    return {
        "title": case.title,
        "objective": case.objective,
        "precondition": case.precondition,
        "postcondition": case.postcondition,
        "related_requirement": case.related_requirement,
        "steps": [{"order": s.order, "action": s.action, "expected": s.expected} for s in case.steps],
        "raw_html": case.raw_html or "",
    }


def _build_review_comment(hard_fails: list[str]) -> str:
    """Build a review comment from failed checklist item IDs."""
    parts: list[str] = []
    for item_id in hard_fails:
        desc = CHECKLIST.get(item_id, (item_id, ""))[0]
        parts.append(f"{item_id}: {desc}")
    return "Previous case failed hard-gate checks. Fix these issues: " + "; ".join(parts)


def _build_req_info_for_eval(
    req: RequirementInput,
    analysis,
    set_meta: dict | None,
) -> dict:
    """Build the req_info dict needed by evaluate_case()."""
    signals = analysis.signals if analysis else []
    thresholds = analysis.thresholds if analysis else []
    timing = [t for t in (analysis.timing if analysis else [])
              if t.strip().lower() != "none found"]

    # Expected missing categories: prefer requirement set, fall back to LLM#1 analysis
    expected_missing: list[str] = []
    if set_meta and "expected_missing_categories" in set_meta:
        expected_missing = set_meta["expected_missing_categories"]
    elif analysis and analysis.missing_info_items:
        expected_missing = [mi.category for mi in analysis.missing_info_items if mi.category]

    return {
        "signals": signals,
        "thresholds": thresholds,
        "timing": timing,
        "case_coverage": "",
        "requirement_description": req.description,
        "supplementary_info": req.supplementary_info,
        "accepted_test_basis": set_meta.get("accepted_test_basis", "") if set_meta else "",
        "expected_missing_categories": expected_missing,
    }


def _self_check_case(case, analysis, provider) -> tuple:
    """Run LLM self-check for invented identifiers.

    Returns (corrected_case, had_changes: bool).
    If the LLM call fails or parsing fails, returns the original case unchanged.
    """
    if not analysis or not (analysis.signals or analysis.observations or analysis.states):
        return case, False

    try:
        known_signals = ", ".join(analysis.signals) if analysis.signals else "(none)"
        known_observations = ", ".join(analysis.observations) if analysis.observations else "(none)"
        known_states = ", ".join(analysis.states) if analysis.states else "(none)"
        case_html = case.raw_html or ""

        sys_prompt, usr_prompt = render_prompt(
            "self_check",
            known_signals=known_signals,
            known_observations=known_observations,
            known_states=known_states,
            case_html=case_html,
        )
        output = provider.complete(sys_prompt, usr_prompt)
        corrected = parse_generated_case(output)
        if corrected and corrected.steps and corrected.title:
            return corrected, True
    except Exception:
        pass

    return case, False


def _cell_str(row: tuple, index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    val = row[index]
    return str(val).strip() if val is not None else ""


def read_excel(
    file_path: str,
    requirement_key_col: str,
    description_col: str,
    type_col: str,
    function_name_col: str,
) -> list[dict]:
    """Read all rows from Excel, returning list with type preserved."""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    wb.close()

    key_idx = headers.index(requirement_key_col) if requirement_key_col in headers else -1
    desc_idx = headers.index(description_col) if description_col in headers else -1
    type_idx = headers.index(type_col) if type_col in headers else -1
    func_idx = headers.index(function_name_col) if function_name_col in headers else -1

    results: list[dict] = []
    for row_idx, row in enumerate(rows, start=2):
        req_type = _cell_str(row, type_idx)
        results.append({
            "requirement_key": _cell_str(row, key_idx),
            "description": _cell_str(row, desc_idx),
            "type": req_type,
            "function_name": _cell_str(row, func_idx),
            "source_row": row_idx,
        })
    return results


def build_requirement_inputs(rows: list[dict]) -> list[RequirementInput]:
    """Convert rows to RequirementInput list, injecting heading/info context.

    Only rows with type='Requirement' produce a RequirementInput.
    Each requirement gets its preceding headings (hierarchy stack) and info rows
    as supplementary context.
    """
    heading_stack: list[str] = []
    pending_info: list[str] = []
    inputs: list[RequirementInput] = []

    for row in rows:
        row_type = row["type"].lower()

        if row_type == "heading":
            # Reset or update heading hierarchy based on heading text
            heading_text = row["description"]
            # Simple hierarchy tracking: if it starts with a lower number, it's a new section
            heading_stack.append(heading_text)
            pending_info.clear()

        elif row_type == "info":
            pending_info.append(row["description"])

        elif row_type in ("requirement", ""):
            # Build context
            context_parts: list[str] = []
            if heading_stack:
                context_parts.append("Section: " + " > ".join(heading_stack))
            if pending_info:
                context_parts.append("Context: " + " | ".join(pending_info))

            # If the requirement has a function_name, note it
            if row.get("function_name"):
                context_parts.insert(0, f"Function: {row['function_name']}")

            inputs.append(RequirementInput(
                requirement_key=row["requirement_key"],
                description=row["description"],
                function_name=row.get("function_name", ""),
                supplementary_info="\n".join(context_parts) if context_parts else "",
            ))

            # Don't clear heading_stack on requirement — same section may have more reqs
            pending_info.clear()

    return inputs


def sample_requirements(
    inputs: list[RequirementInput],
    sample_size: int,
    seed: int | None = None,
) -> list[RequirementInput]:
    """Randomly sample requirements."""
    if seed is not None:
        random.seed(seed)
    if len(inputs) <= sample_size:
        return inputs[:]
    return random.sample(inputs, sample_size)


# ── Requirement set loading ──────────────────────────────────────────────


def load_requirement_set(path: str) -> dict:
    """Load and validate a requirement set JSON file.

    Returns the parsed dict on success. Raises ValueError with a
    human-readable message on failure.
    """
    set_path = Path(path)
    if not set_path.exists():
        raise ValueError(f"Requirement set file not found: {path}")

    try:
        data = json.loads(set_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Requirement set file is not valid JSON: {exc}") from exc

    validate_requirement_set(data, path)
    return data


def validate_requirement_set(data: dict, path: str) -> None:
    """Validate a loaded requirement set dict in-place. Raises ValueError."""
    if not isinstance(data, dict):
        raise ValueError(f"Requirement set must be a JSON object, got {type(data).__name__}")

    name = data.get("name")
    if not isinstance(name, str) or not name.strip():
        raise ValueError("Requirement set is missing a non-empty 'name' field")

    entries = data.get("entries")
    if not isinstance(entries, list) or len(entries) == 0:
        raise ValueError(f"Requirement set '{name}' has no 'entries' list")

    keys_seen: set[str] = set()
    for i, entry in enumerate(entries):
        if not isinstance(entry, dict):
            raise ValueError(f"Entry {i} in '{name}' is not a JSON object")

        key = entry.get("requirement_key")
        if not isinstance(key, str) or not key.strip():
            raise ValueError(f"Entry {i} in '{name}' is missing a valid 'requirement_key'")

        if key in keys_seen:
            raise ValueError(f"Duplicate requirement_key '{key}' at entry {i} in '{name}'")
        keys_seen.add(key)

        bucket = entry.get("evaluation_bucket")
        if not isinstance(bucket, str) or not bucket.strip():
            raise ValueError(f"Entry '{key}' in '{name}' is missing 'evaluation_bucket'")

        cats = entry.get("expected_missing_categories")
        if not isinstance(cats, list):
            raise ValueError(
                f"Entry '{key}' in '{name}': 'expected_missing_categories' must be a list"
            )
        invalid = [c for c in cats if c not in _VALID_CATEGORIES]
        if invalid:
            raise ValueError(
                f"Entry '{key}' in '{name}': invalid expected_missing_categories {invalid}. "
                f"Allowed: {sorted(_VALID_CATEGORIES)}"
            )

        # Validate inline content (only required for self-contained sets)
        desc = entry.get("description", "")
        if not isinstance(desc, str) or not desc.strip():
            raise ValueError(
                f"Entry '{key}' in '{name}' is missing 'description'. "
                f"Requirement sets must be self-contained with inline content."
            )


def build_requirements_from_set(req_set_data: dict) -> list[RequirementInput]:
    """Build RequirementInput list directly from a self-contained requirement set.

    Each entry must have: requirement_key, description.
    Optional: function_name, supplementary_info.
    """
    selected: list[RequirementInput] = []
    for entry in req_set_data["entries"]:
        selected.append(RequirementInput(
            requirement_key=entry["requirement_key"],
            description=entry["description"],
            function_name=entry.get("function_name", ""),
            supplementary_info=entry.get("supplementary_info", ""),
        ))
    return selected


def select_by_requirement_set(
    all_inputs: list[RequirementInput],
    req_set_data: dict,
) -> list[RequirementInput]:
    """Select and order RequirementInputs by a requirement set from Excel.

    Preserves the order in the set. Raises ValueError if any requirement
    key in the set is not found in all_inputs.

    Only used when --excel is provided alongside --requirement-set.
    """
    lookup: dict[str, RequirementInput] = {
        ri.requirement_key: ri for ri in all_inputs
    }
    entries = req_set_data["entries"]
    selected: list[RequirementInput] = []
    missing: list[str] = []

    for entry in entries:
        key = entry["requirement_key"]
        req = lookup.get(key)
        if req is None:
            missing.append(key)
        else:
            selected.append(req)

    if missing:
        raise ValueError(
            f"{len(missing)} requirement key(s) from the set not found in Excel: "
            + ", ".join(missing)
        )

    return selected


# ── Core pipeline ────────────────────────────────────────────────────────


def archive_prompts(round_dir: Path) -> None:
    """Copy current review-pipeline prompt files into round_dir/prompts/."""
    prompts_dir = round_dir / "prompts"
    prompts_dir.mkdir(parents=True, exist_ok=True)

    project_root = Path(__file__).resolve().parents[1]
    source_dir = project_root / "src" / "testcase_agent" / "review_pipeline" / "prompts"

    for src in sorted(source_dir.glob("*.html")):
        # Strip .html extension for the archived copy.
        dest_name = src.stem + ".md"
        shutil.copy2(src, prompts_dir / dest_name)
        print(f"  Archived prompt: {src.name} -> {dest_name}")


def _serialize_case_for_output(
    case,
    report,
    *,
    retry_meta: dict | None = None,
    sanitize_enabled: bool = False,
    sanitize_replacements: list[str] | None = None,
) -> dict:
    """Serialize one generated case for generated_cases.json."""
    replacements = sanitize_replacements or []
    out: dict = {
        "title": case.title,
        "objective": case.objective,
        "precondition": case.precondition,
        "postcondition": case.postcondition,
        "related_requirement": case.related_requirement,
        "steps": [
            {"order": s.order, "action": s.action, "expected": s.expected}
            for s in case.steps
        ],
        "raw_html": case.raw_html,
        "sanitize": {
            "enabled": sanitize_enabled,
            "replacement_count": len(replacements),
            "replacements": replacements,
        },
        "quality": {
            "passed": report.passed,
            "failures": report.failures,
            "warnings": report.warnings,
        },
    }
    if retry_meta:
        out["retry"] = retry_meta
    return out


def run_batch(
    requirements: list[RequirementInput],
    output_dir: Path,
    sanitize: bool = True,
    requirement_set_data: dict | None = None,
    run_eval: bool = False,
) -> dict:
    """Legacy batch generation entry point.

    Kept only as an importable guard for older helper code. New generation must
    start from ``python -m testcase_agent.review_pipeline.cli prepare-clarification-review``.
    """
    settings = get_settings()
    provider = create_provider(settings)

    output_dir.mkdir(parents=True, exist_ok=True)
    archive_prompts(output_dir)

    # Build per-key lookup from the set if provided
    set_lookup: dict[str, dict] = {}
    if requirement_set_data:
        for entry in requirement_set_data["entries"]:
            set_lookup[entry["requirement_key"]] = entry

    all_results: list[dict] = []
    total_cases = 0
    errors: list[dict] = []

    # ── Background DeepSeek evaluator setup ──────────────────────────
    eval_executor = None
    eval_futures: dict = {}
    results_by_idx: dict[int, list] = {}
    _eval_errors: list[int] = [0]
    _eval_buffer: list[tuple[int, dict]] = []
    _eval_flush = None  # type: ignore[assignment]
    if run_eval:
        from concurrent.futures import ThreadPoolExecutor
        from optimization.claude_evaluator import (
            build_system_prompt, score_batch, save_evaluation,
            _print_requirement_score,
            EvalResult, DEFAULT_MODEL as EVAL_DEFAULT_MODEL,
        )

        eval_system_prompt = build_system_prompt()
        _BATCH_SIZE = 5
        eval_executor = ThreadPoolExecutor(max_workers=3)
        eval_total = len(requirements)

        def _submit_batch(batch_entries: list[tuple[int, dict]]):
            entries = [e for _, e in batch_entries]
            indices = [i for i, _ in batch_entries]

            def _score_batch_entries():
                try:
                    scores = score_batch(
                        entries, eval_system_prompt, EVAL_DEFAULT_MODEL,
                        start_idx=indices[0] + 1, total=eval_total,
                    )
                    return indices, scores, None
                except Exception as exc:
                    return indices, [], exc

            future = eval_executor.submit(_score_batch_entries)
            future.add_done_callback(_on_batch_done)
            eval_futures[future] = indices

        def _on_batch_done(future):
            indices = eval_futures[future]
            try:
                _, scores, exc = future.result()
                if exc:
                    for idx in indices:
                        req_key = all_results[idx].get("requirement_key", f"#{idx}")
                        print(f"  DeepSeek [{idx + 1}/{eval_total}] {req_key} FAILED: {exc}")
                        _eval_errors[0] += 1
                else:
                    score_by_key = {s.requirement_key: s for s in scores}
                    for idx in indices:
                        entry = all_results[idx]
                        req_key = entry.get("requirement_key", f"#{idx}")
                        s = score_by_key.get(req_key)
                        if s:
                            results_by_idx[idx] = [s]
                            _print_requirement_score(idx, eval_total, req_key, s)
                        else:
                            _print_requirement_score(idx, eval_total, req_key, None)
                            _eval_errors[0] += 1
            except Exception as exc:
                for idx in indices:
                    req_key = all_results[idx].get("requirement_key", f"#{idx}")
                    print(f"  DeepSeek [{idx + 1}/{eval_total}] {req_key} FAILED: {exc}")
                    _eval_errors[0] += 1
            sys.stdout.flush()

        def _eval_flush():
            if _eval_buffer:
                _submit_batch(list(_eval_buffer))
                _eval_buffer.clear()

    for i, req in enumerate(requirements):
        print(f"[{i+1}/{len(requirements)}] Generating: {req.requirement_key} ...")
        result = run_pipeline(req, provider)
        gen_failures = set(result.generation_failures)

        if result.error:
            errors.append({
                "requirement_key": req.requirement_key,
                "error": result.error,
            })
            print(f"  ERROR: {result.error}")
            continue

        # ── Retry + self-check ───────────────────────────────────────
        set_meta = set_lookup.get(req.requirement_key)
        req_info = _build_req_info_for_eval(req, result.analysis, set_meta)
        intents = result.analysis.case_intents if result.analysis else []
        sanitize_enabled_for_case = bool(sanitize and result.analysis)
        sigs = result.analysis.signals if result.analysis else []
        thr = result.analysis.thresholds if result.analysis else []
        tim = [
            t for t in (result.analysis.timing if result.analysis else [])
            if t.strip().lower() != "none found"
        ]

        def _post_process_case(case, retry_meta: dict) -> tuple:
            corrected, sc_changed = _self_check_case(case, result.analysis, provider)
            if sc_changed:
                case = corrected
                retry_meta["self_check_changed"] = True

            replacements: list[str] = []
            if sanitize_enabled_for_case:
                case, replacements = sanitize_numeric_values(
                    case,
                    requirement_description=req.description,
                    supplementary_info=req.supplementary_info,
                    extracted_signals=sigs,
                    extracted_thresholds=thr,
                    extracted_timing=tim,
                    accepted_test_basis=req_info.get("accepted_test_basis", ""),
                )
            return case, replacements

        retry_metas: list[dict] = []
        sanitize_replacements_by_case: list[list[str]] = []
        new_cases: list = []
        for ci, case in enumerate(result.cases):
            intent = intents[ci] if ci < len(intents) else None
            retry_meta: dict = {"attempts": 0, "exhausted": False, "failures": [], "self_check_changed": False, "generation_timeout": ci in gen_failures}
            sanitize_replacements: list[str] = []

            # Step 1: Self-check identifiers, then sanitize invented numeric values.
            case, replacements = _post_process_case(case, retry_meta)
            sanitize_replacements.extend(replacements)

            # Step 2: Retry loop (max 2 retries + initial = 3 attempts)
            for attempt in range(3):
                hard_fails, _ = evaluate_case(_case_to_dict(case), req_info, {})
                if not hard_fails:
                    break
                retry_meta["failures"].append(hard_fails)
                if attempt < 2 and intent:
                    review = "" if retry_meta["generation_timeout"] else _build_review_comment(hard_fails)
                    try:
                        case = regenerate_case(
                            req, intent.description, intent.coverage,
                            review, provider, analysis=result.analysis,
                        )
                        retry_meta["attempts"] = attempt + 1
                        case, replacements = _post_process_case(case, retry_meta)
                        sanitize_replacements.extend(replacements)
                    except Exception:
                        retry_meta["exhausted"] = True
                        break
                else:
                    retry_meta["exhausted"] = True

            new_cases.append(case)
            retry_metas.append(retry_meta)
            sanitize_replacements_by_case.append(sanitize_replacements)

        result.cases = new_cases
        total_sanitize_replacements = sum(len(reps) for reps in sanitize_replacements_by_case)
        if total_sanitize_replacements:
            print(f"  sanitize: {total_sanitize_replacements} value(s) replaced with [NEEDS REVIEW]")

        # ── Quality gate (runtime schema) ────────────────────────────
        quality_reports = evaluate_cases(result.cases)

        analysis_data = None
        if result.analysis:
            analysis_data = {
                "signals": result.analysis.signals,
                "thresholds": result.analysis.thresholds,
                "timing": result.analysis.timing,
                "states": result.analysis.states,
                "observations": result.analysis.observations,
                "direction": result.analysis.direction,
                "missing_critical_info": result.analysis.missing_critical_info,
                "missing_info_items": [
                    {"category": mi.category, "description": mi.description}
                    for mi in result.analysis.missing_info_items
                ],
                "case_intents": [
                    {"coverage": ci.coverage, "description": ci.description}
                    for ci in result.analysis.case_intents
                ],
            }

        cases_data = []
        for case, report, retry_meta, replacements in zip(
            result.cases,
            quality_reports,
            retry_metas,
            sanitize_replacements_by_case,
        ):
            cases_data.append(_serialize_case_for_output(
                case,
                report,
                retry_meta=retry_meta,
                sanitize_enabled=sanitize_enabled_for_case,
                sanitize_replacements=replacements,
            ))

        entry: dict = {
            "requirement_key": req.requirement_key,
            "function_name": req.function_name,
            "description": req.description,
            "supplementary_info": req.supplementary_info,
            "analysis": analysis_data,
            "cases": cases_data,
        }
        # Enrich with requirement set metadata when available
        if set_meta:
            entry["evaluation_bucket"] = set_meta["evaluation_bucket"]
            entry["expected_missing_categories"] = set_meta["expected_missing_categories"]
            entry["requirement_set_note"] = set_meta.get("rationale", "")
        all_results.append(entry)
        total_cases += len(cases_data)

        passed_count = sum(1 for r in quality_reports if r.passed)
        retried = sum(1 for rm in retry_metas if rm["attempts"] > 0)
        print(f"  {len(cases_data)} cases, quality: {passed_count}/{len(cases_data)} passed, {retried} retried")

        # ── Buffer entries for batched DeepSeek scoring ────────────────
        if eval_executor is not None:
            _eval_buffer.append((i, entry))
            if len(_eval_buffer) >= _BATCH_SIZE:
                _eval_flush()

    # Flush remaining DeepSeek scoring batches
    if run_eval:
        _eval_flush()

    # Save generated cases
    cases_path = output_dir / "generated_cases.json"
    cases_path.write_text(
        json.dumps(all_results, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Save sampled requirements list
    req_list = [
        {
            "requirement_key": r.requirement_key,
            "description": r.description,
            "function_name": r.function_name,
        }
        for r in requirements
    ]
    req_path = output_dir / "sampled_requirements.json"
    req_path.write_text(
        json.dumps(req_list, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    summary: dict = {
        "total_requirements": len(requirements),
        "total_cases": total_cases,
        "errors": len(errors),
        "error_details": errors,
    }
    if requirement_set_data:
        summary["requirement_set_name"] = requirement_set_data.get("name", "")
        summary["requirement_set_path"] = str(
            requirement_set_data.get("_source_path", "")
        )
        summary["total_requirement_set_entries"] = len(
            requirement_set_data.get("entries", [])
        )
    summary_path = output_dir / "summary.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Save hard-rule evaluation
    try:
        from optimization.evaluator import evaluate_generated_cases, save_evaluation_result
        with open(cases_path, encoding="utf-8") as f:
            gen_data = json.load(f)
        hr_result = evaluate_generated_cases(gen_data)
        save_evaluation_result(hr_result, "hardrule", output_dir)
        print(f"Hard-rule evaluation saved ({hr_result.case_pass_rate}%)")
    except Exception as exc:
        print(f"Hard-rule evaluation save failed: {exc}")

    # Wait for background DeepSeek evaluation and save results
    if run_eval:
        try:
            eval_executor.shutdown(wait=True)
            eval_result = EvalResult(model_used=EVAL_DEFAULT_MODEL)
            for i in sorted(results_by_idx.keys()):
                eval_result.requirements.extend(results_by_idx[i])
            eval_result.errors = _eval_errors[0]
            save_evaluation(eval_result, output_dir, evaluator_name="deepseek")
            print(f"DeepSeek evaluation completed (weighted={eval_result.overall_weighted})")
        except Exception as exc:
            print(f"DeepSeek evaluation failed: {exc}")
            if eval_executor is not None:
                eval_executor.shutdown(wait=False)

    # Generate unified cases_report.html
    try:
        from optimization.generate_case_html import generate_round_html
        generate_round_html(output_dir, 1)
        report_path = output_dir / "cases_report.html"
        print(f"Report: {report_path}")
    except Exception as exc:
        print(f"Report generation failed: {exc}")

    print(f"\nDone. {len(requirements)} requirements → {total_cases} cases, {len(errors)} errors")
    print(f"Output: {output_dir}")

    return summary


def main():
    parser = argparse.ArgumentParser(description="Testcase Agent optimization CLI")
    sub = parser.add_subparsers(dest="command", required=True)

    # run command
    run_parser = sub.add_parser("run", help="Generate cases for sampled requirements")
    run_parser.add_argument("--excel", default=None, help="Path to Excel requirements file (optional if --requirement-set provides inline content)")
    run_parser.add_argument("--output-dir", required=True, help="Output directory for this round")
    run_parser.add_argument("--sample", type=int, default=20, help="Number of requirements to sample (default: 20, ignored with --requirement-set)")
    run_parser.add_argument("--seed", type=int, default=None, help="Random seed for sampling (for reproducibility)")
    run_parser.add_argument("--requirement-set", default=None, help="Path to a requirement set JSON file (e.g. prompt_eval_v1.json)")
    run_parser.add_argument("--key-col", default="Requirement ID")
    run_parser.add_argument("--desc-col", default="Requirement Description")
    run_parser.add_argument("--type-col", default="Type")
    run_parser.add_argument("--func-col", default="function")
    run_parser.add_argument("--limit", type=int, default=None, help="Limit to first N requirements (for quick testing, works with --requirement-set)")
    run_parser.add_argument("--no-sanitize", action="store_true", help="Disable post-processing that replaces invented numeric values with [NEEDS REVIEW] (sanitize is ON by default)")
    run_parser.add_argument("--eval", action="store_true", help="Run DeepSeek 8-dimension evaluation after generation and generate complete report")

    # evaluate command
    eval_parser = sub.add_parser("evaluate", help="Run AI evaluation against checklist_v2.md")
    eval_parser.add_argument("--round-dir", required=True, help="Path to a round directory containing generated_cases.json")
    eval_parser.add_argument("--model", default=None, help="Model ID (default: from ANTHROPIC_MODEL env or deepseek-v4-flash[1m])")
    eval_parser.add_argument("--delay", type=float, default=0.1, help="Seconds between API submissions (default: 0.1)")
    eval_parser.add_argument("--concurrency", type=int, default=5, help="Max parallel scoring calls (default: 5)")
    eval_parser.add_argument("--report", action="store_true", help="Regenerate cases_report.html from existing evaluation files")

    args = parser.parse_args()

    if args.command == "run":
        parser.error(
            "optimization.cli run was removed with the legacy generation "
            "pipeline. Use python -m testcase_agent.review_pipeline.cli "
            "prepare-clarification-review instead."
        )

    elif args.command == "evaluate":
        round_dir = Path(args.round_dir)
        if not round_dir.is_absolute():
            round_dir = _PROJECT_ROOT / round_dir

        if args.report:
            from optimization.generate_case_html import generate_round_html
            generate_round_html(round_dir, 1)
        else:
            from optimization.claude_evaluator import DEFAULT_MODEL, run_full_evaluation
            run_full_evaluation(
                round_dir,
                model=args.model or DEFAULT_MODEL,
                delay=args.delay,
                max_concurrency=args.concurrency,
            )


if __name__ == "__main__":
    main()

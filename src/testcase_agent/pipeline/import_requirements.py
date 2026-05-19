from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import load_workbook


@dataclass
class ColumnMapping:
    requirement_key_col: str = ""
    description_col: str = ""
    function_name_col: str = ""
    requirement_type_col: str = ""
    supplementary_info_cols: list[str] = field(default_factory=list)


def list_columns(file_path: str, sheet_name: str | None = None) -> list[str]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [str(cell.value) if cell.value is not None else f"Column_{i}" for i, cell in enumerate(ws[1], start=1)]
    wb.close()
    return headers


def list_sheets(file_path: str) -> list[str]:
    wb = load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


@dataclass
class ParsedRequirement:
    requirement_key: str
    description: str
    function_name: str = ""
    requirement_type: str = ""
    supplementary_info: str = ""
    source_row: int = 0

    @property
    def is_heading(self) -> bool:
        return self.requirement_type.lower() == "heading"

    @property
    def is_info(self) -> bool:
        return self.requirement_type.lower() == "info"

    @property
    def is_requirement(self) -> bool:
        return not self.is_heading and not self.is_info


def parse_requirements(
    file_path: str,
    mapping: ColumnMapping,
    sheet_name: str | None = None,
) -> list[ParsedRequirement]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    headers = list_columns(file_path, sheet_name)

    key_idx = _col_index(headers, mapping.requirement_key_col)
    desc_idx = _col_index(headers, mapping.description_col)
    func_idx = _col_index(headers, mapping.function_name_col)
    type_idx = _col_index(headers, mapping.requirement_type_col)
    supp_indices = [_col_index(headers, col) for col in mapping.supplementary_info_cols if col]

    requirements: list[ParsedRequirement] = []
    for row_idx, row in enumerate(rows, start=2):
        key = _cell_str(row, key_idx)
        desc = _cell_str(row, desc_idx)
        if not key or not desc:
            continue
        func = _cell_str(row, func_idx)
        req_type = _cell_str(row, type_idx)

        supp_parts: list[str] = []
        for si in supp_indices:
            val = _cell_str(row, si)
            if val:
                supp_parts.append(val)
        supp = " | ".join(supp_parts)

        requirements.append(ParsedRequirement(
            requirement_key=key,
            description=desc,
            function_name=func,
            requirement_type=req_type,
            supplementary_info=supp,
            source_row=row_idx,
        ))

    return requirements


def _col_index(headers: list[str], col_name: str) -> int:
    try:
        return headers.index(col_name)
    except ValueError:
        return -1


def _cell_str(row: tuple, index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    val = row[index]
    return str(val).strip() if val is not None else ""
